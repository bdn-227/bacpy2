
# import libraries
import polars as pl
from os.path import basename
from numpy import intersect1d, arange
from numpy.char import strip
from re import findall
from itertools import chain
from numpy import argwhere, array
from openpyxl import load_workbook
import numpy as np


def convert_long_chunk_to_values(chunk, used_wells, excel_path):
    """
    function that takes a long-formatted chunk and gets rid of all the empty cells

    """
    # extract the values
    values = chunk.filter(chunk[:,0].is_in(used_wells))
    #first, try to convert column to nulls and remove
    for column in chunk.columns:
        try:
            values = values.with_columns(pl.col(column).cast(pl.Null))
            if values[column].is_null().all():
                values = values.select(pl.exclude(column))
        except:
            pass
            #raise ValueError(f"ask Nik to fix this stuff here..\nuse this to debug {excel_path}")
    # reformat the columns
    for column in values.columns:
        if values[column].dtype == pl.String:
            values = values.with_columns(pl.col(column).replace("OVER", 1e10))
        try:
            values = values.with_columns(pl.col(column).cast(float))
        except:
            pass
    # get measurement mode
    measurement_mode = chunk.filter(pl.col('column_0').str.starts_with("Mode"))['column_4'][0]
    # parse the metadata depending on wheter this is a scan measurement or single
    if values.shape[1] > 2: 
        # branch for scan
        measurement_type = "scan"
        # reformat values-df
        ex = int(chunk.filter(pl.col("column_0") == "Excitation Wavelength").select("column_4")[0,0])
        em = chunk.filter(pl.col("column_0") == "Wavel.").to_numpy().reshape(-1)
        em = em[em != None]
        em = [incom for incom in em if str(incom) != 'nan']
        values.columns = array(em)
        values = values.rename({"Wavel.": "well"}).unpivot(index="well", variable_name="em", value_name="response").with_columns(pl.col("em").cast(float).cast(int))
    else: 
        # branch for single
        measurement_type = "single"
        # reformat the values-df
        values.columns = ["well", "response"]
        if measurement_mode == 'Fluorescence Bottom Reading':
            excitation = chunk['Tecan i-control , 2.0.10.0'].loc[chunk.iloc[:, 0].str.startswith("Excitation Wavelength", na=False)].iloc[0]
            emission = chunk['Tecan i-control , 2.0.10.0'].loc[chunk.iloc[:, 0].str.startswith("Emission Wavelength", na=False)].iloc[0]
            raise ValueError(f"ask Nik to fix the parser for single fluorescent bottom reading..\nuse this to debug {excel_path}")
        else:
            ex = int(chunk.filter(pl.col("column_0").str.starts_with("Measurement Wavelength")).select("column_4")[0,0])
            em = ex
        # add emission to 
        values = values.with_columns(pl.lit(em).alias("em"))
    # add emission information
    values = (values
                .with_columns(pl.lit(ex).alias("ex"))
                .with_columns(pl.lit(measurement_type).alias("measurement_type"))
                .with_columns(pl.lit(measurement_mode).alias("measurement_mode"))
                .with_columns(pl.col("response").cast(float))
             )
    return values


def convert_chunk(chunk, all_wells, excel_path, spark=False, idx=0):
    #chunk = chunks[23]
    #print(idx)
    # get the used wells
    used_wells = intersect1d(chunk[:,0].drop_nulls(), all_wells)
    # determine mode of operation (TECAN outputs can have different formats, long or wide)
    if len(used_wells) == 0:
        # get the performed measurement modes
        modes = chunk.filter(pl.col("column_0").str.contains("Mode")).to_numpy()
        # branch for OD in wide format
        if "Absorbance" in modes:
            wv = list(chunk.filter(pl.col("column_0").str.contains("(?i)Measurement Wavelength")).to_numpy().reshape(-1))
            wv = int([val for val in wv if str(val).isdigit()][0])
            measurement_idx = argwhere(chunk["column_0"].is_in(["A", "B", "C", "D", "E", "F", "G", "H"])).reshape(-1)
            column_row_idx = measurement_idx.min() -1 
            values = chunk[measurement_idx]
            columns_dict = {key: value for key, value in chunk[int(column_row_idx)].to_dict().items() if value.is_not_null().all()}
            values = values.select(columns_dict)
            values.columns = [str(val[0]) for val in columns_dict.values()]
            values = values.unpivot(index="<>")
            values.columns = ["row", "column", "response"]
            values = values.with_columns( (pl.col("row") + pl.col("column").cast(pl.Float32).cast(int).cast(str)).alias("well"),
                                          (pl.lit(wv).alias("em")), 
                                          (pl.lit(wv).alias("ex")), 
                                          (pl.lit("single").alias("measurement_type")), 
                                          (pl.lit("Absorbance").alias("measurement_mode")),
                                          (pl.col("response").cast(float))
                                        ).select("well", "response", "em", "ex", "measurement_type", "measurement_mode").drop_nulls()
            # final list of columns
            # │ well ┆ response ┆ em  ┆ ex  ┆ measurement_type ┆ measurement_mode
            # │ str  ┆ f64      ┆ i32 ┆ i32 ┆ str              ┆ str
        if 'Fluorescence Intensity Scan Bottom Reading' in modes:
            measurement_idx = argwhere(chunk["column_0"].is_in(["A", "B", "C", "D", "E", "F", "G", "H"])).reshape(-1)
            if len(measurement_idx) == 0:
                measurement_idx = argwhere(chunk["column_0"].str.starts_with("Wavel."))[0,0]
                ex = int(chunk.filter(pl.col("column_0").str.starts_with("Excitation wavelength [nm]"))[0,4])
                values = chunk[measurement_idx:]
                new_cols = [col for col in list(values[0].to_numpy()[0]) if col is not None]
                values = values[:, range(len(new_cols))]
                values.columns = new_cols
                values = values.filter(~pl.col("Wavel.").str.starts_with("Wavel.")).drop_nulls()
                values = (values
                                .unpivot(index="Wavel.", variable_name="well", value_name="response")
                                .rename({"Wavel.": "em"})
                                .with_columns(pl.lit('Fluorescence Bottom Reading').alias("measurement_mode"))
                                .with_columns(pl.lit('scan' if values.shape[0] > 1 else "single").alias("measurement_type"))
                                .with_columns(pl.lit(ex).alias("ex"))
                                .select("well", "response", "em", "ex", "measurement_type", "measurement_mode")
                                .with_columns(pl.col("response").replace("OVER", 1e10).cast(float))
                                )
            else:
                raise ValueError(f"ask Nik to fix the parser for wide-format fluorescent data processing..\nuse this to debug {excel_path}")
    else:
        # for long format
        values = convert_long_chunk_to_values(chunk, used_wells, excel_path)
    # yield error if non-complete measurement
    if values.shape[0] == 0:
        print(f"incomplete measurement in this file:\n{excel_path}")
    else:
        # now extract the complete metadata
        if spark:
            date = "-".join(chunk.filter(pl.col("column_0").str.starts_with("Start Time"))[0,4].split(" ")[0].split(".")[::-1])
            start_time = chunk.filter(pl.col("column_0").str.starts_with("Start Time"))[0,4].split(" ")[1]
            label = chunk.filter(pl.col("column_0").str.starts_with("Name")).select("column_1")[0,0]
            temp = float(chunk.filter(pl.col("column_0").str.starts_with("Temperature"))["column_4"][0])
        else:
            date = "-".join(chunk.filter(pl.col("column_0").str.starts_with("Start Time:"))[0,1].split(" ")[0].split(".")[::-1])
            start_time = chunk.filter(pl.col("column_0").str.starts_with("Start Time:"))[0,1].split(" ")[1]
            label = chunk.filter(pl.col("column_0").str.starts_with("Label")).select("column_0")[0,0].split(" ")[1]
            temp = float(chunk.filter(pl.col("column_1").str.starts_with("Temperature: "))["column_1"][0].split(" ")[1])
        # put data together        
        values = (values
                        .with_columns(pl.lit(start_time).alias("start_time"))
                        .with_columns(pl.lit(label).alias("label"))
                        .with_columns(pl.lit(temp).alias("temp"))
                        .with_columns(pl.lit(date).alias("date"))
                        )
        return values.select(['well', 'ex', 'em', 'response', 'measurement_type', 'measurement_mode', 'start_time', 'label', 'temp', 'date'])



def get_kinetic_metadata(raw):
    meta_start = raw[:,0].str.contains("Labels", literal=True).arg_max()
    meta_end = raw[:,0].str.contains("Start Time:", literal=True).arg_max()
    metadata = raw[meta_start:meta_end+1,:]
    block_starts = metadata[:,0].str.starts_with("Mode").arg_true()
    block_ends = (block_starts[1:]).to_list() + metadata[:,0].str.contains("Start Time:", literal=True).arg_true().to_list()
    start_time = metadata[-1,1]
    date = start_time.split(" ")[0]
    start_time = start_time.split(" ")[1]
    measurement_type = "kinetic"
    meta_ls = []
    for s, e in zip(block_starts, block_ends):
        m = metadata[s:e,:]
        idx = (m[:,0] == "Mode").arg_max()
        measurement_mode = m[idx, "column_4"]
        if measurement_mode == 'Fluorescence Bottom Reading':
            idx = (m[:,0] == "Excitation Wavelength").arg_max()
            ex = int(m[idx, "column_4"])
            idx = (m[:,0] == "Emission Wavelength").arg_max()
            em = int(m[idx, "column_4"])
        if measurement_mode == 'Absorbance':
            idx = (m[:,0] == "Measurement Wavelength").arg_max()
            ex = int(m[idx, "column_4"])
            em = int(m[idx, "column_4"])
        
        meta_res = pl.DataFrame({"measurement_type": measurement_type,
                                "start_time": start_time,
                                "date": date,
                                "measurement_mode": measurement_mode,
                                "ex": ex,
                                "em": em,
                                })
        meta_ls.append(meta_res)
    return meta_ls



def parse_kinetic_chunk(chunk, all_wells, metadata):
    label = chunk[0,0]
    chunk = chunk.rename(chunk.head(2).to_dicts().pop())[2:,:]
    chunk = chunk.unpivot(on=np.intersect1d(chunk.columns, all_wells), index=[col for col in chunk.columns if col not in all_wells], value_name="response", variable_name="well")
    column_mapping = {"Cycle Nr.": "cycle_n", "Time [s]": "time_s", "Temp. [°C]": "temp"}
    chunk = chunk.rename(column_mapping)
    chunk = (chunk
                .join(metadata, how="cross")
                .with_columns(pl.lit(label).alias("label"))
                .with_columns(pl.col("cycle_n").cast(int))
                .with_columns(pl.col("time_s").cast(float))
                .with_columns(pl.col("temp").cast(float))
                .with_columns(pl.col("response").cast(float))
                .with_columns(pl.col("start_time").str.strptime(pl.Time, "%H:%M:%S"))
                .with_columns(pl.col("date").str.strptime(pl.Date, "%d.%m.%Y").alias("date"))
                )
    return chunk.select(['well', 'ex', 'em', 'response', 'measurement_type', 'measurement_mode', 'start_time', 'label', 'temp', 'date', "cycle_n", "time_s"])



def parse_file_icontrol(excel_path):
    """
    function to parse excel files containg spectra information obtained using the
    TECAN iControl software
    excel_path:     str     path to an excel file created by TECAN iControl
    """
    # first, determine which books to read
    workbook = load_workbook(excel_path, read_only=True)
    sheet_names = workbook.sheetnames
    print(f"parsing: {sheet_names}")
    sheet_ls = []
    for sheet in sheet_names:
        # read the input file
        try:
            raw = pl.read_excel(excel_path, sheet_name=sheet, engine="xlsx2csv", infer_schema_length=1000000)
        except:
            print(f"skipping empty sheet: {sheet}")
            continue
        raw = raw.rename({old: f"column_{idx}" for idx, old in enumerate(raw.columns)})
        # define all possible wells for a typical 96-well assay
        cols = arange(1, 25)
        rows = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', "I", "J", "K", "L", "M", "N", "O", "P"]
        all_wells = list(chain.from_iterable([[i+str(j) for j in cols] for i in rows]))
        # conditional branch here
        spark = "SparkControl" in raw[0,0]
        if spark:
            """
            if true, we have a spark script
            """
            device = int(findall(r'\d+', raw["column_4"][1])[0])

            # split the dataframe into chunks from "LabelXX" to "End Time"
            starts = raw[:,0].str.starts_with("Mode").arg_true()
            ends   = raw[:,0].str.starts_with("End Time").arg_true()
            kinetic = False ######## <------- in case a  'NoneType' object has no attribute 'split error occurs, here is where 
            # the correct logic goes
        else:
            """
            normal tecan icontrol script
            """
            device = int(findall(r'\d+', raw["column_4"][0])[0])
            # determine if kinetic cycle
            kinetic_idx = raw[:,0].str.contains("Kinetic Cycles", literal=True).arg_max()
            kinetic = kinetic_idx and int(raw[kinetic_idx,4]) >= 2
            if kinetic:
                starts = raw[:,0].str.starts_with("Label").arg_true()
                ends   = (starts[1:]-1).to_list() + (raw[:,0].str.starts_with("End Time").arg_true()-1).to_list()
            else:
                starts = raw[:,0].str.starts_with("Label").arg_true()
                ends   = raw[:,0].str.starts_with("End Time").arg_true()
        #create a list with subsets of df
        chunks = [raw[start:end+1] for start, end in zip(starts, ends)]
        if kinetic:
            kinetic_metadata_ls = get_kinetic_metadata(raw)
            parsed_ls = [parse_kinetic_chunk(chunk, all_wells, meta) for chunk, meta in zip(chunks, kinetic_metadata_ls)]
            parsed_ls = [df for df in parsed_ls if df is not None]
            parsed_chunks = pl.concat(parsed_ls, how="vertical_relaxed")
        # branch for a "normal", fluorescent type of measurement
        else:
            parsed_ls = [convert_chunk(chunk, all_wells, excel_path, spark, idx) for idx, chunk in enumerate(chunks)]
            parsed_ls = [df for df in parsed_ls if df is not None]
            parsed_chunks = pl.concat(parsed_ls, how="vertical_relaxed")
        # now add the residual columns
        keep_cols = [
                    'filename', 
                    'date', 
                    'device', 
                    'measurement_mode', 
                    'ex', 
                    'em', 
                    'temp',
                    'start_time',
                    'label', 
                    'measurement_type', 
                    'row', 
                    'column',
                    'well', 
                    'ex_em',
                    'response',
                    "sheet",
                    ]
        parsed_chunks = (parsed_chunks
                            .with_columns(pl.lit(basename(excel_path)).alias("filename"))
                            .with_columns(pl.lit(device).alias("device"))
                            .with_columns(pl.Series(strip(parsed_chunks["well"].to_numpy().astype(str), "|".join(cols.astype(str)))).alias("row"))
                            .with_columns(pl.Series(strip(parsed_chunks["well"].to_numpy().astype(str), "|".join(rows))).cast(int).alias("column"))
                            .with_columns(("wv" + pl.col("ex").cast(str) + "." + pl.col("em").cast(str)).alias("ex_em"))
                            .with_columns(pl.lit(sheet).alias("sheet"))
                            .with_columns( (pl.col("filename") + "_" + pl.col("sheet")).alias("filename") )
                            .select(keep_cols + ["cycle_n", "time_s"] if kinetic else keep_cols)
                            )
        sheet_ls.append(parsed_chunks)
    return pl.concat(sheet_ls, how="vertical_relaxed")